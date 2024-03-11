import pytest
import os
import datetime
from openpyxl.reader.excel import load_workbook

from sqlalchemy import create_engine, func
from sqlalchemy.orm import Session

from jkrimporter import conf
from jkrimporter.cli.jkr import import_lopetusilmoitukset
from jkrimporter.providers.db.database import json_dumps
from jkrimporter.providers.db.models import Kompostori
from jkrimporter.providers.lahti.ilmoitustiedosto import Ilmoitustiedosto


@pytest.fixture(scope="module", autouse=True)
def engine():
    engine = create_engine(
        "postgresql://{username}:{password}@{host}:{port}/{dbname}".format(
            **conf.dbconf
        ),
        future=True,
        json_serializer=json_dumps,
    )
    return engine


def test_readable(datadir):
    assert Ilmoitustiedosto.readable_by_me(datadir + "/lopetusilmoitukset.xlsx")


def test_lopetusilmoitus(engine, datadir):
    import_lopetusilmoitukset(datadir + "/lopetusilmoitukset.xlsx")
    session = Session(engine)
    end_date = datetime.date(2022, 8, 18)
    # Kahdelle kompostorille asettuu loppupäivämääräksi 18.8.2022.
    assert session.query(func.count(Kompostori.id)).filter(Kompostori.loppupvm == end_date).scalar() == 2

    # Etsitään tiedosto jonka nimi sisältää "kohdentumattomat_ilmoitus".
    files_in_dir = os.listdir(datadir)
    matching_files = [
        filename for filename in files_in_dir if "kohdentumattomat_lopetusilmoitukset" in filename
    ]
    # Kohdentumattomat_ilmoitus löytyy.
    assert len(matching_files) == 1

    # Kohdentumattomat tiedostossa kaksi riviä.
    xlsx_file_path = os.path.join(datadir, matching_files[0])
    workbook = load_workbook(xlsx_file_path)
    sheet = workbook[workbook.sheetnames[0]]
    assert sheet.max_row == 2