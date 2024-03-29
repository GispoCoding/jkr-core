import logging
import os
from pathlib import Path

from openpyxl.reader.excel import load_workbook
from pydantic import ValidationError

from jkrimporter.datasheets import get_paatostiedosto_headers
from jkrimporter.providers.lahti.models import Paatos
from jkrimporter.utils.paatos import export_kohdentumattomat_paatokset

logger = logging.getLogger(__name__)


class Paatostiedosto:
    def __init__(self, path):
        self._path = path

    @classmethod
    def readable_by_me(cls, path):
        file = Path(path)
        if file.is_file() and file.suffix == ".xlsx":
            try:
                workbook = load_workbook(filename=file, data_only=True, read_only=True)
                if len(workbook.sheetnames):
                    return True
            except Exception:
                pass
        return False

    @property
    def paatokset(self):
        paatos_list = []
        missing_headers_list = []
        failed_validations = []

        workbook = load_workbook(self._path)
        sheet = workbook[workbook.sheetnames[0]]

        # In Excel files, the row indices start from 1.
        headers = [cell.value for cell in sheet[1]]
        for header in get_paatostiedosto_headers():
            if header not in headers:
                missing_headers_list.append(header)

        if missing_headers_list:
            print(
                f"Tiedosto: {self._path}, puuttuvat sarakeotsikot: {missing_headers_list}"
            )
            raise RuntimeError("Päätöstiedostosta puuttuu oletettuja sarakeotsikoita.")

        for row in sheet.iter_rows(min_row=2, values_only=True):
            try:
                data = dict(zip(headers, row))
                paatos_obj = Paatos.parse_obj(data)
                paatos_obj.rawdata = data
                paatos_list.append(paatos_obj)
            except ValidationError as e:
                logger.warning(
                    f"Paatos-olion luonti epäonnistui datalla: {row}. Virhe: {e}"
                )
                failed_validations.append(data)

        if failed_validations:
            print(f"Päätostiedosto sisältää {len(failed_validations)} virheellistä riviä.")
            export_kohdentumattomat_paatokset(os.path.dirname(self._path), failed_validations)

        return paatos_list
